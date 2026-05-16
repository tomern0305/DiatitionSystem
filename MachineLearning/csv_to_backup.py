"""
Convert daily_food_nutrition_dataset.csv → dataset_backup.zip
compatible with /api/system/import.

Fetches one food image per product from Pixabay (rate-limited to 1 req/sec).
Requires PIXABAY_KEY in MachineLearning/.env

Usage:
    cd MachineLearning
    python csv_to_backup.py
Output:
    MachineLearning/dataset_backup.zip
"""
import json, zipfile, io, pathlib, time, os
import requests
import pandas as pd
from openpyxl import Workbook
from dotenv import load_dotenv

ROOT     = pathlib.Path(__file__).resolve().parent
CSV_PATH = ROOT / "daily_food_nutrition_dataset.csv"
OUT_PATH = ROOT / "dataset_backup.zip"

load_dotenv(ROOT / ".env")
PIXABAY_KEY = os.getenv("PIXABAY_KEY", "").strip()

# Cache: search term → image URL (avoids duplicate API calls for similar names)
_image_cache: dict[str, str | None] = {}

def fetch_image(food_name: str) -> str | None:
    """Search Pixabay for a food photo. Returns the web-format URL or None."""
    # Strip serving-size annotations like "(2 large)" or "(1 slice)" before searching
    import re
    clean = re.sub(r"\(.*?\)", "", food_name).strip()
    query = f"{clean} food"
    if query in _image_cache:
        return _image_cache[query]
    try:
        resp = requests.get(
            "https://pixabay.com/api/",
            params={"key": PIXABAY_KEY, "q": query, "image_type": "photo",
                    "category": "food", "per_page": 3, "safesearch": "true"},
            timeout=10,
        )
        hits = resp.json().get("hits", []) if resp.ok else []
        url = hits[0]["webformatURL"] if hits else None
    except Exception:
        url = None
    _image_cache[query] = url
    return url

df = pd.read_csv(CSV_PATH, on_bad_lines="skip")

# Build categories from unique values in the CSV (1-indexed)
categories = sorted(df["Category"].dropna().unique())
cat_map = {name: i + 1 for i, name in enumerate(categories)}

wb = Workbook()

# ── Categories sheet ──────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Categories"
ws.append(["id", "name"])
for name, cid in cat_map.items():
    ws.append([cid, name])

# ── Empty lookup sheets (no data from CSV) ────────────────────────────────────
for title in ("Sensitivities", "Textures", "Diets"):
    s = wb.create_sheet(title)
    s.append(["id", "name"])

# ── Products sheet ────────────────────────────────────────────────────────────
ps = wb.create_sheet("Products")
ps.append([
    "id", "name", "category_id", "image_url", "iddsi",
    "calories", "protein", "carbs", "fat", "sugars", "sodium",
    "contains", "may_contain", "texture_id", "properties",
    "company", "texture_notes", "allergy_notes", "forbidden_for",
    "nutrition_vector", "openai_embedding",
])
total = len(df)
for i, row in df.iterrows():
    cal  = float(row.get("Calories (kcal)", 0) or 0)
    prot = float(row.get("Protein (g)", 0) or 0)
    carb = float(row.get("Carbohydrates (g)", 0) or 0)
    fat  = float(row.get("Fat (g)", 0) or 0)
    sug  = float(row.get("Sugars (g)", 0) or 0)
    sod  = float(row.get("Sodium (mg)", 0) or 0)
    vec  = json.dumps([cal, prot, carb, fat, sug, sod])

    food_name = str(row["Food_Item"])
    img_url   = fetch_image(food_name) if PIXABAY_KEY else None
    # 1 req/sec — Pixabay free tier allows ~100/min, this stays well under
    time.sleep(1)

    print(f"  [{i+1}/{total}] {food_name[:40]:<40}  {'✓' if img_url else '–'}")

    ps.append([
        i + 1,                               # id
        food_name,                           # name (translate to Hebrew post-import)
        cat_map.get(row.get("Category"), None),  # category_id
        img_url,                             # image_url from Pixabay
        None,                                # iddsi
        cal, prot, carb, fat, sug, sod,
        "[]", "[]",                          # contains / may_contain
        None,                                # texture_id
        "[]",                                # properties
        None, None, None, None,              # company, texture_notes, allergy_notes, forbidden_for
        vec,                                 # nutrition_vector (enables similarity search)
        None,                                # openai_embedding
    ])

# ── Empty Meals sheet ─────────────────────────────────────────────────────────
ms = wb.create_sheet("Meals")
ms.append([
    "id", "name", "description", "diet_id", "product_ids",
    "nutrition", "filter_restriction_ids", "filter_texture_ids",
    "filter_show_may_contain", "created_at",
])

# ── Pack into ZIP ─────────────────────────────────────────────────────────────
xlsx_buf = io.BytesIO()
wb.save(xlsx_buf)
xlsx_buf.seek(0)

with zipfile.ZipFile(OUT_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
    zf.writestr("backup.xlsx", xlsx_buf.read())

print(f"Written: {OUT_PATH}  ({OUT_PATH.stat().st_size // 1024} KB)")
print(f"Products: {len(df)}   Categories: {len(cat_map)}")
